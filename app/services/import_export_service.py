from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Optional

from lxml import etree
from openpyxl import Workbook, load_workbook

from app.config import VALID_DOMAINS, VALID_PARTS_OF_SPEECH
from app.database import db
from app.models import TermCreate, new_id, now
from app.services.audit_service import add_audit_log
from app.services.term_service import create_term


def _idempotent_key(source_term: str, source_lang: str, target_lang: str, domain: str) -> str:
    """Build a deterministic key used for import deduplication."""
    return f"{source_term}|{source_lang}|{target_lang}|{domain}"


def _find_existing_term(key: str) -> Optional[dict]:
    """Find an existing term whose idempotent key matches."""
    for term in db.terms.values():
        tkey = _idempotent_key(
            term["source_term"], term["source_lang"], term["target_lang"], term["domain"]
        )
        if tkey == key:
            return term
    return None


def _upsert_term(data: TermCreate, user_id: str) -> str:
    """Create a new term or increment version if an entry with the same idempotent key exists."""
    key = _idempotent_key(data.source_term, data.source_lang, data.target_lang, data.domain)
    existing = _find_existing_term(key)
    if existing:
        if data.domain not in VALID_DOMAINS:
            raise ValueError(f"Invalid domain: {data.domain}")
        if data.part_of_speech not in VALID_PARTS_OF_SPEECH:
            raise ValueError(f"Invalid part_of_speech: {data.part_of_speech}")
        existing["version"] += 1
        existing["source_term"] = data.source_term
        existing["source_lang"] = data.source_lang
        existing["target_term"] = data.target_term
        existing["target_lang"] = data.target_lang
        existing["domain"] = data.domain
        existing["definition"] = data.definition
        existing["example_source"] = data.example_source
        existing["example_target"] = data.example_target
        existing["forbidden_terms"] = list(data.forbidden_terms)
        existing["synonyms"] = list(data.synonyms)
        existing["case_sensitive"] = data.case_sensitive
        existing["part_of_speech"] = data.part_of_speech
        existing["is_forced"] = data.is_forced
        existing["updated_at"] = now().isoformat()

        add_audit_log(
            term_id=existing["term_id"],
            action="import_update",
            from_status=existing["status"],
            to_status=existing["status"],
            from_version=existing["version"] - 1,
            to_version=existing["version"],
            changed_by=user_id,
            reason="Term updated via import (idempotent key match)",
            diff=None,
        )

        db.fts_update(
            existing["term_id"],
            existing["source_term"],
            existing["target_term"],
            existing.get("definition", ""),
            existing["domain"],
            existing["source_lang"],
            existing["target_lang"],
        )
        return "updated"

    create_term(data, user_id)
    return "created"


def import_tbx(content: bytes, user_id: str) -> dict:
    """Parse TBX (TermBase eXchange) XML and import term entries with idempotent upsert logic."""
    root = etree.fromstring(content)
    created = 0
    updated = 0

    body = root.find(".//{urn:iso:std:iso:30042:ed-2}body")
    if body is None:
        body = root.find(".//body")
    if body is None:
        return {"created": 0, "updated": 0}

    for concept_entry in body.iterchildren():
        tag = etree.QName(concept_entry).localname if isinstance(concept_entry.tag, str) else ""
        if tag != "conceptEntry":
            continue

        source_term = ""
        target_term = ""
        source_lang = ""
        target_lang = ""
        domain = ""
        definition = ""

        for lang_sec in concept_entry.iterchildren():
            lang_tag = etree.QName(lang_sec).localname if isinstance(lang_sec.tag, str) else ""
            if lang_tag != "langSec":
                continue

            lang_code = lang_sec.get("{urn:iso:std:iso:30042:ed-2}lang", "")
            if not lang_code:
                lang_code = lang_sec.get("lang", "")

            for term_sec in lang_sec.iterchildren():
                sec_tag = etree.QName(term_sec).localname if isinstance(term_sec.tag, str) else ""
                if sec_tag == "termSec":
                    term_el = term_sec.find("{urn:iso:std:iso:30042:ed-2}term")
                    if term_el is None:
                        term_el = term_sec.find("term")
                    if term_el is not None and term_el.text:
                        if not source_term:
                            source_term = term_el.text.strip()
                            source_lang = lang_code
                        else:
                            target_term = term_el.text.strip()
                            target_lang = lang_code

                elif sec_tag == "descripSec":
                    descrip = term_sec.find("{urn:iso:std:iso:30042:ed-2}descrip")
                    if descrip is None:
                        descrip = term_sec.find("descrip")
                    if descrip is not None:
                        dtype = descrip.get("type", "")
                        if dtype == "domain" and descrip.text:
                            domain = descrip.text.strip()
                        elif dtype == "definition" and descrip.text:
                            definition = descrip.text.strip()

        if not source_term or not target_term:
            continue
        if domain and domain not in VALID_DOMAINS:
            domain = "通用"

        data = TermCreate(
            source_term=source_term,
            source_lang=source_lang or "en",
            target_term=target_term,
            target_lang=target_lang or "zh",
            domain=domain or "通用",
            definition=definition,
        )
        result = _upsert_term(data, user_id)
        if result == "created":
            created += 1
        else:
            updated += 1

    return {"created": created, "updated": updated}


_CSV_HEADERS = [
    "source_term", "source_lang", "target_term", "target_lang",
    "domain", "definition", "example_source", "example_target",
    "forbidden_terms", "synonyms", "case_sensitive", "part_of_speech", "is_forced",
]


def import_csv(content: bytes, user_id: str) -> dict:
    """Parse CSV bytes and import term entries with idempotent upsert logic."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    created = 0
    updated = 0

    for row in reader:
        source_term = row.get("source_term", "").strip()
        target_term = row.get("target_term", "").strip()
        if not source_term or not target_term:
            continue

        domain = row.get("domain", "").strip() or "通用"
        if domain not in VALID_DOMAINS:
            domain = "通用"

        forbidden_raw = row.get("forbidden_terms", "").strip()
        forbidden_terms = [t.strip() for t in forbidden_raw.split(";") if t.strip()] if forbidden_raw else []

        synonyms_raw = row.get("synonyms", "").strip()
        synonyms = [t.strip() for t in synonyms_raw.split(";") if t.strip()] if synonyms_raw else []

        case_sensitive_raw = row.get("case_sensitive", "false").strip().lower()
        case_sensitive = case_sensitive_raw == "true"

        is_forced_raw = row.get("is_forced", "true").strip().lower()
        is_forced = is_forced_raw != "false"

        data = TermCreate(
            source_term=source_term,
            source_lang=row.get("source_lang", "en").strip() or "en",
            target_term=target_term,
            target_lang=row.get("target_lang", "zh").strip() or "zh",
            domain=domain,
            definition=row.get("definition", "").strip(),
            example_source=row.get("example_source", "").strip(),
            example_target=row.get("example_target", "").strip(),
            forbidden_terms=forbidden_terms,
            synonyms=synonyms,
            case_sensitive=case_sensitive,
            part_of_speech=row.get("part_of_speech", "noun").strip() or "noun",
            is_forced=is_forced,
        )
        result = _upsert_term(data, user_id)
        if result == "created":
            created += 1
        else:
            updated += 1

    return {"created": created, "updated": updated}


def import_excel(content: bytes, user_id: str) -> dict:
    """Parse Excel (.xlsx) bytes and import term entries with idempotent upsert logic."""
    wb = load_workbook(io.BytesIO(content), read_only=True)
    created = 0
    updated = 0

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            continue
        headers = [str(h).strip() if h else "" for h in header_row]

        for row in rows:
            row_dict = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = str(val) if val is not None else ""

            source_term = row_dict.get("source_term", "").strip()
            target_term = row_dict.get("target_term", "").strip()
            if not source_term or not target_term:
                continue

            domain = row_dict.get("domain", "").strip() or "通用"
            if domain not in VALID_DOMAINS:
                domain = "通用"

            forbidden_raw = row_dict.get("forbidden_terms", "").strip()
            forbidden_terms = [t.strip() for t in forbidden_raw.split(";") if t.strip()] if forbidden_raw else []

            synonyms_raw = row_dict.get("synonyms", "").strip()
            synonyms = [t.strip() for t in synonyms_raw.split(";") if t.strip()] if synonyms_raw else []

            case_sensitive_raw = row_dict.get("case_sensitive", "false").strip().lower()
            case_sensitive = case_sensitive_raw == "true"

            is_forced_raw = row_dict.get("is_forced", "true").strip().lower()
            is_forced = is_forced_raw != "false"

            data = TermCreate(
                source_term=source_term,
                source_lang=row_dict.get("source_lang", "en").strip() or "en",
                target_term=target_term,
                target_lang=row_dict.get("target_lang", "zh").strip() or "zh",
                domain=domain,
                definition=row_dict.get("definition", "").strip(),
                example_source=row_dict.get("example_source", "").strip(),
                example_target=row_dict.get("example_target", "").strip(),
                forbidden_terms=forbidden_terms,
                synonyms=synonyms,
                case_sensitive=case_sensitive,
                part_of_speech=row_dict.get("part_of_speech", "noun").strip() or "noun",
                is_forced=is_forced,
            )
            result = _upsert_term(data, user_id)
            if result == "created":
                created += 1
            else:
                updated += 1

    wb.close()
    return {"created": created, "updated": updated}


def export_tbx(terms: list[dict]) -> bytes:
    """Export terms to TBX XML format, grouped by domain."""
    nsmap = {"xml": "http://www.w3.org/XML/1998/namespace"}
    tbx = etree.Element("tbx", nsmap=nsmap)
    body = etree.SubElement(tbx, "body")

    grouped: dict[str, list[dict]] = {}
    for term in terms:
        d = term.get("domain", "通用")
        grouped.setdefault(d, []).append(term)

    for domain, domain_terms in grouped.items():
        for term in domain_terms:
            concept_entry = etree.SubElement(body, "conceptEntry")

            lang_sec_src = etree.SubElement(concept_entry, "langSec")
            lang_sec_src.set("lang", term.get("source_lang", "en"))
            term_sec_src = etree.SubElement(lang_sec_src, "termSec")
            term_el_src = etree.SubElement(term_sec_src, "term")
            term_el_src.text = term.get("source_term", "")

            if term.get("definition"):
                descrip_sec = etree.SubElement(concept_entry, "descripSec")
                descrip = etree.SubElement(descrip_sec, "descrip")
                descrip.set("type", "definition")
                descrip.text = term["definition"]

            if domain:
                descrip_sec_d = etree.SubElement(concept_entry, "descripSec")
                descrip_d = etree.SubElement(descrip_sec_d, "descrip")
                descrip_d.set("type", "domain")
                descrip_d.text = domain

            lang_sec_tgt = etree.SubElement(concept_entry, "langSec")
            lang_sec_tgt.set("lang", term.get("target_lang", "zh"))
            term_sec_tgt = etree.SubElement(lang_sec_tgt, "termSec")
            term_el_tgt = etree.SubElement(term_sec_tgt, "term")
            term_el_tgt.text = term.get("target_term", "")

    return etree.tostring(tbx, xml_declaration=True, encoding="UTF-8", pretty_print=True)


def export_csv(terms: list[dict]) -> bytes:
    """Export terms to CSV bytes with standard headers."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_CSV_HEADERS, extrasaction="ignore")
    writer.writeheader()

    for term in terms:
        row = {
            "source_term": term.get("source_term", ""),
            "source_lang": term.get("source_lang", "en"),
            "target_term": term.get("target_term", ""),
            "target_lang": term.get("target_lang", "zh"),
            "domain": term.get("domain", ""),
            "definition": term.get("definition", ""),
            "example_source": term.get("example_source", ""),
            "example_target": term.get("example_target", ""),
            "forbidden_terms": ";".join(term.get("forbidden_terms", [])),
            "synonyms": ";".join(term.get("synonyms", [])),
            "case_sensitive": str(term.get("case_sensitive", False)).lower(),
            "part_of_speech": term.get("part_of_speech", "noun"),
            "is_forced": str(term.get("is_forced", True)).lower(),
        }
        writer.writerow(row)

    return buf.getvalue().encode("utf-8-sig")


def export_excel(terms: list[dict]) -> bytes:
    """Export terms to Excel (.xlsx) bytes with one sheet per domain."""
    wb = Workbook(write_only=True)

    grouped: dict[str, list[dict]] = {}
    for term in terms:
        d = term.get("domain", "通用")
        grouped.setdefault(d, []).append(term)

    for domain, domain_terms in grouped.items():
        ws = wb.create_sheet(title=domain[:31])
        ws.append(_CSV_HEADERS)

        for term in domain_terms:
            ws.append([
                term.get("source_term", ""),
                term.get("source_lang", "en"),
                term.get("target_term", ""),
                term.get("target_lang", "zh"),
                term.get("domain", ""),
                term.get("definition", ""),
                term.get("example_source", ""),
                term.get("example_target", ""),
                ";".join(term.get("forbidden_terms", [])),
                ";".join(term.get("synonyms", [])),
                str(term.get("case_sensitive", False)).lower(),
                term.get("part_of_speech", "noun"),
                str(term.get("is_forced", True)).lower(),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_filtered_terms_for_export(
    status: Optional[str] = None,
    domain: Optional[str] = None,
    approved_by: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> list[dict]:
    """Filter terms from db.terms based on export criteria and return matching dicts."""
    results = []
    for term in db.terms.values():
        if status and term.get("status") != status:
            continue
        if domain and term.get("domain") != domain:
            continue
        if approved_by and term.get("approved_by") != approved_by:
            continue
        if date_from:
            created_at = term.get("created_at", "")
            if created_at < date_from:
                continue
        if date_to:
            created_at = term.get("created_at", "")
            if created_at > date_to:
                continue
        results.append(term)
    return results
